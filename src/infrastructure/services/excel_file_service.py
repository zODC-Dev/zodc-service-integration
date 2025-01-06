from io import BytesIO
from typing import Any, Dict

from fastapi import UploadFile
import openpyxl

from src.domain.exceptions.excel_file_exceptions import ExcelFileInvalidFormatError, ExcelFileReadError
from src.domain.services.excel_file_service import IExcelFileService


class ExcelFileService(IExcelFileService):
    async def extract_file(self, file: UploadFile) -> Dict[str, Any]:
        # Validate file type
        if not file.filename.endswith((".xlsx", ".xls")):
            raise ExcelFileInvalidFormatError(
                "Invalid file format. Please upload an Excel file.") from None

        # Read file content
        contents = await file.read()
        try:
            workbook = openpyxl.load_workbook(
                filename=BytesIO(contents), data_only=True)
        except Exception as e:
            raise ExcelFileReadError("Fail to load excel file") from e

        # Process each sheet
        result: Dict[str, Any] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract header
            headers = [cell.value for cell in sheet[1]]

            sheet_result: Dict[str, Any] = {header: [] for header in headers}

            # Extract all content in 1 sheet
            for row in sheet.iter_rows(min_row=2):
                for cell, header in zip(row, headers):
                    sheet_result[header].append(
                        str(cell.value) if cell.value is not None else "")

            if not result:
                result = sheet_result
            else:
                for key, value in sheet_result.items():
                    if key in result:
                        result[key].extend(value)
                    else:
                        result[key] = value

        return result
